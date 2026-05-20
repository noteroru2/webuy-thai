from __future__ import annotations

import json
import re
from pathlib import Path
from urllib.parse import urlparse, unquote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = Path(r"C:\Users\User\Downloads\xn--c3c3a0aa6cvaf8b9dze.com-Coverage-Valid-2026-05-20.xlsx")
POSTS_DIR = ROOT / "src" / "content" / "posts"
PAGES_DIR = ROOT / "src" / "pages"


def normalize_path(path: str) -> str:
    if not path:
        return "/"
    path = unquote(path).strip()
    if not path.startswith("/"):
        path = "/" + path
    path = re.sub(r"/{2,}", "/", path)
    if path != "/" and not path.endswith("/"):
        path += "/"
    return path


def extract_frontmatter_lines(text: str) -> list[str]:
    if text.startswith("\ufeff"):
        text = text[1:]
    if not text.startswith("---"):
        return []
    end = text.find("\n---", 3)
    if end == -1:
        return []
    return text[4:end].splitlines()


def parse_frontmatter(md_path: Path) -> dict[str, str | bool]:
    data: dict[str, str | bool] = {
        "slug": "",
        "redirectTo": "",
        "noindex": False,
        "gone": False,
        "title": "",
    }
    lines = extract_frontmatter_lines(md_path.read_text(encoding="utf-8"))
    for line in lines:
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key in {"slug", "redirectTo", "title"}:
            data[key] = value
        if key in {"noindex", "gone"}:
            data[key] = value == "true"
    return data


def page_file_to_route(page_path: Path) -> str | None:
    rel = page_path.relative_to(PAGES_DIR).as_posix()
    if rel in {"[slug].astro", "404.astro"}:
        return None
    if rel.endswith("/index.astro"):
        route = "/" + rel[: -len("/index.astro")] + "/"
    elif rel.endswith(".astro"):
        route = "/" + rel[: -len(".astro")] + "/"
    elif rel.endswith(".ts"):
        route = "/" + rel[: -len(".ts")]
        if not route.endswith(".txt"):
            route += "/"
        return route
    else:
        return None
    return normalize_path(route)


def classify_keyword(path: str, title: str = "") -> dict[str, str]:
    haystack = f"{path} {title}".lower()
    province_rules = [
        ("อุบลราชธานี", ["อุบลราชธานี", "อุบล"]),
        ("กรุงเทพมหานคร", ["กรุงเทพมหานคร", "กรุงเทพ", "krungthep", "bangkok"]),
        ("ขอนแก่น", ["ขอนแก่น", "khon-kaen", "khonkaen"]),
        ("นครราชสีมา", ["นครราชสีมา", "โคราช", "korat", "nakhon-ratchasima"]),
        ("อุดรธานี", ["อุดรธานี", "udon"]),
        ("ชลบุรี", ["ชลบุรี", "chonburi", "chon-buri"]),
        ("ยโสธร", ["ยโสธร", "yasothon"]),
        ("ร้อยเอ็ด", ["ร้อยเอ็ด", "roi-et"]),
        ("มหาสารคาม", ["มหาสารคาม", "mahasarakham", "สารคาม"]),
        ("อำนาจเจริญ", ["อำนาจเจริญ"]),
        ("ศรีสะเกษ", ["ศรีสะเกษ"]),
        ("นครพนม", ["นครพนม"]),
    ]
    product_rules = [
        ("โน๊ตบุ๊ค", ["โน๊ตบุ๊ค", "notebook", "laptop"]),
        ("คอมพิวเตอร์", ["คอม", "คอมพิวเตอร์", "computer", "pc"]),
        ("iPhone", ["ไอโฟน", "iphone"]),
        ("iPad", ["ไอแพด", "ipad"]),
        ("MacBook", ["แมคบุ๊ค", "macbook"]),
        ("กล้อง", ["กล้อง", "camera", "canon", "nikon", "fujifilm", "sony", "olympus", "โดรน", "drone"]),
        ("ลำโพง", ["ลำโพง", "marshall", "jbl"]),
        ("Android", ["android", "samsung", "oppo", "vivo", "xiaomi", "มือถือ", "โทรศัพท์"]),
        ("Apple Watch", ["apple-watch", "apple watch"]),
        ("เครื่องเกม", ["ps4", "ps5", "play station", "เครื่องเกม"]),
    ]
    intent_rules = [
        ("รับซื้อ", ["รับซื้อ", "rab-sue"]),
        ("รับจำนำ", ["รับจำนำ", "จำนำ", "ตั๋วจำนำ"]),
        ("ขาย", ["ขาย", "ร้านขาย"]),
        ("ซ่อม", ["ซ่อม"]),
        ("คู่มือ", ["คู่มือ", "เช็ก", "ราคา", "ขายได้ไหม"]),
    ]

    province = "ไม่ระบุ"
    product = "ทั่วไป"
    intent = "ทั่วไป"

    for label, keywords in province_rules:
        if any(k in haystack for k in keywords):
            province = label
            break
    for label, keywords in product_rules:
        if any(k in haystack for k in keywords):
            product = label
            break
    for label, keywords in intent_rules:
        if any(k in haystack for k in keywords):
            intent = label
            break
    return {"province": province, "product": product, "intent": intent}


def importance_score(path: str, title: str = "") -> int:
    parts = classify_keyword(path, title)
    score = 0
    if parts["intent"] == "รับซื้อ":
        score += 4
    elif parts["intent"] == "คู่มือ":
        score += 2
    elif parts["intent"] == "ขาย":
        score += 1
    elif parts["intent"] == "รับจำนำ":
        score -= 2

    if parts["product"] in {"โน๊ตบุ๊ค", "คอมพิวเตอร์", "iPhone", "iPad", "MacBook", "กล้อง", "Android"}:
        score += 3
    elif parts["product"] in {"ลำโพง", "เครื่องเกม"}:
        score += 1
    elif parts["product"] == "Apple Watch":
        score -= 1

    if parts["province"] in {"อุบลราชธานี", "กรุงเทพมหานคร", "ขอนแก่น", "นครราชสีมา", "อุดรธานี", "ชลบุรี"}:
        score += 3
    elif parts["province"] in {"ยโสธร", "ร้อยเอ็ด", "มหาสารคาม", "อำนาจเจริญ", "ศรีสะเกษ"}:
        score += 2
    elif parts["province"] != "ไม่ระบุ":
        score += 1

    if any(token in path for token in ["ใกล้ฉัน", "มือสอง", "ราค", "ขายได้"]):
        score += 1

    return score


def main() -> None:
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["ตาราง"]

    coverage_paths = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[0]
        if not url:
            continue
        parsed = urlparse(str(url))
        coverage_paths.append(normalize_path(parsed.path))

    static_routes = set()
    for page in PAGES_DIR.rglob("*"):
        if page.is_file():
            route = page_file_to_route(page)
            if route:
                static_routes.add(route)

    posts = []
    post_routes = set()
    redirects = set()
    gone = set()
    noindex = set()
    title_map: dict[str, str] = {}

    for md_path in POSTS_DIR.rglob("*.md"):
        data = parse_frontmatter(md_path)
        slug = str(data["slug"]).strip()
        if not slug:
            continue
        route = normalize_path(slug)
        post_routes.add(route)
        title_map[route] = str(data.get("title", "")).strip()
        posts.append(
            {
                "route": route,
                "title": str(data.get("title", "")).strip(),
                "redirectTo": str(data.get("redirectTo", "")).strip(),
                "noindex": bool(data.get("noindex", False)),
                "gone": bool(data.get("gone", False)),
            }
        )
        if data.get("redirectTo"):
            redirects.add(route)
        if data.get("gone"):
            gone.add(route)
        if data.get("noindex"):
            noindex.add(route)

    current_routes = static_routes | post_routes

    exact_present = []
    present_redirect = []
    present_gone = []
    missing = []

    for route in coverage_paths:
        title = title_map.get(route, "")
        row = {
            "route": route,
            "title": title,
            "classification": classify_keyword(route, title),
            "importance": importance_score(route, title),
        }
        if route in redirects:
            present_redirect.append(row)
        elif route in gone:
            present_gone.append(row)
        elif route in current_routes:
            exact_present.append(row)
        else:
            missing.append(row)

    missing.sort(key=lambda item: (-item["importance"], item["route"]))

    report = {
        "coverage_total": len(coverage_paths),
        "exact_present": len(exact_present),
        "present_redirect": len(present_redirect),
        "present_gone": len(present_gone),
        "missing": len(missing),
        "top_missing": missing[:60],
    }

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
